import os
import json
import io
import random
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from models import (
    init_db, get_db, check_admin, get_setting, set_setting, DB_PATH,
    get_project_mode, set_project_mode
)

app = Flask(__name__)
app.secret_key = "practice-select-fixed-key-2024-kayson"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

init_db()

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

# ========== 定时检查（per-project）==========
def check_time_based_open(project_id):
    """定时开放：仅在timer_enabled=1时自动切换，不影响手动操作"""
    conn = get_db()
    row = conn.execute("SELECT timer_enabled, open_time, close_time, status FROM projects WHERE id=?", (project_id,)).fetchone()
    conn.close()
    if not row or not row["timer_enabled"]:
        return
    open_time = row["open_time"]
    close_time = row["close_time"]
    if not open_time or not close_time:
        return
    try:
        now = datetime.now()
        open_dt = datetime.strptime(open_time, "%Y-%m-%dT%H:%M")
        close_dt = datetime.strptime(close_time, "%Y-%m-%dT%H:%M")
        status = row["status"]
        if open_dt <= now <= close_dt:
            if status in ("close", "preview"):
                set_project_mode(project_id, "open")
        elif now > close_dt:
            if status in ("open", "preview"):
                set_project_mode(project_id, "close")
    except:
        pass

# ========== Admin Auth ==========
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "").strip()
        if check_admin(u, p):
            session["admin"] = True
            return redirect(url_for("admin_projects"))
        flash("账号或密码错误", "danger")
    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin_login"))

# ========== 项目管理（新增）==========
@app.route("/admin")
@admin_required
def admin_projects():
    """项目列表首页"""
    conn = get_db()
    projects = conn.execute("""
        SELECT p.*,
               (SELECT COUNT(*) FROM students s WHERE s.project_id=p.id) as student_count,
               (SELECT COUNT(*) FROM positions pos WHERE pos.project_id=p.id) as position_count,
               (SELECT COUNT(*) FROM submissions sub WHERE sub.project_id=p.id) as submitted_count
        FROM projects p ORDER BY p.id DESC
    """).fetchall()
    conn.close()
    return render_template("admin_projects.html", projects=projects)

@app.route("/admin/project/new", methods=["POST"])
@admin_required
def admin_project_new():
    """新建项目"""
    name = request.form.get("name", "").strip()
    desc = request.form.get("description", "").strip()
    if not name:
        flash("项目名称不能为空", "danger")
        return redirect(url_for("admin_projects"))
    conn = get_db()
    conn.execute("INSERT INTO projects (name, description) VALUES (?, ?)", (name, desc))
    conn.commit()
    conn.close()
    flash(f"项目「{name}」创建成功", "success")
    return redirect(url_for("admin_projects"))

@app.route("/admin/project/<int:pid>/edit", methods=["POST"])
@admin_required
def admin_project_edit(pid):
    """编辑项目"""
    name = request.form.get("name", "").strip()
    desc = request.form.get("description", "").strip()
    if not name:
        flash("项目名称不能为空", "danger")
        return redirect(url_for("admin_projects"))
    conn = get_db()
    conn.execute("UPDATE projects SET name=?, description=? WHERE id=?", (name, desc, pid))
    conn.commit()
    conn.close()
    flash("项目已更新", "success")
    return redirect(url_for("admin_projects"))

@app.route("/admin/project/<int:pid>/delete")
@admin_required
def admin_project_delete(pid):
    """删除项目（级联删除所有数据）"""
    if pid == 1:
        flash("不能删除默认项目", "danger")
        return redirect(url_for("admin_projects"))
    conn = get_db()
    # 级联删除
    conn.execute("DELETE FROM submissions WHERE project_id=?", (pid,))
    conn.execute("DELETE FROM student_position_assign WHERE project_id=?", (pid,))
    conn.execute("DELETE FROM position_classes WHERE position_id IN (SELECT id FROM positions WHERE project_id=?)", (pid,))
    conn.execute("DELETE FROM positions WHERE project_id=?", (pid,))
    conn.execute("DELETE FROM students WHERE project_id=?", (pid,))
    conn.execute("DELETE FROM classes WHERE project_id=?", (pid,))
    conn.execute("DELETE FROM settings WHERE project_id=?", (pid,))
    conn.execute("DELETE FROM projects WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    flash("项目已删除", "warning")
    return redirect(url_for("admin_projects"))

# ========== 项目看板 ==========
@app.route("/admin/project/<int:pid>")
@admin_required
def admin_dashboard(pid):
    conn = get_db()
    project = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    if not project:
        conn.close()
        flash("项目不存在", "danger")
        return redirect(url_for("admin_projects"))
    stats = {}
    stats["total_students"] = conn.execute("SELECT COUNT(*) FROM students WHERE project_id=?", (pid,)).fetchone()[0]
    stats["submitted"] = conn.execute("SELECT COUNT(*) FROM students WHERE project_id=? AND has_submitted=1", (pid,)).fetchone()[0]
    stats["total_positions"] = conn.execute("SELECT COUNT(*) FROM positions WHERE project_id=?", (pid,)).fetchone()[0]
    stats["classes"] = conn.execute("SELECT COUNT(*) FROM classes WHERE project_id=?", (pid,)).fetchone()[0]
    positions = conn.execute("""
        SELECT p.*, 
               GROUP_CONCAT(c.name) as class_names,
               GROUP_CONCAT(pc.class_id) as visible_class_ids
        FROM positions p
        LEFT JOIN position_classes pc ON p.id = pc.position_id
        LEFT JOIN classes c ON pc.class_id = c.id
        WHERE p.project_id = ?
        GROUP BY p.id ORDER BY p.id
    """, (pid,)).fetchall()
    conn.close()
    return render_template("admin_dashboard.html", stats=stats, positions=positions,
                           project=project, pid=pid)

# ========== 系统控制 ==========
@app.route("/admin/project/<int:pid>/system", methods=["POST"])
@admin_required
def admin_system(pid):
    action = request.form.get("action")
    if action == "open":
        set_project_mode(pid, "open")
        set_setting("timer_enabled", "0", project_id=pid)
        flash("系统已开放（学生可选择岗位）", "success")
    elif action == "preview":
        set_project_mode(pid, "preview")
        set_setting("timer_enabled", "0", project_id=pid)
        flash("系统已进入预览模式", "info")
    elif action == "close":
        set_project_mode(pid, "close")
        flash("系统已关闭", "warning")
    elif action == "reset":
        conn = get_db()
        conn.execute("DELETE FROM submissions WHERE project_id=?", (pid,))
        conn.execute("UPDATE students SET has_submitted=0, submitted_at=NULL, assigned_position_id=NULL WHERE project_id=?", (pid,))
        conn.execute("UPDATE positions SET current_count=0 WHERE project_id=?", (pid,))
        conn.commit()
        conn.close()
        flash("已重置所有选择数据", "warning")
    elif action == "set_random":
        random_order = request.form.get("random_order", "0")
        set_setting("random_order", random_order, project_id=pid)
        flash(f"岗位排序已设置为：{'随机展示' if random_order == '1' else '顺序展示'}", "success")
    return redirect(url_for("admin_dashboard", pid=pid))

@app.route("/admin/project/<int:pid>/set_schedule", methods=["POST"])
@admin_required
def admin_set_schedule(pid):
    data = request.get_json()
    open_time = data.get("open_time", "")
    close_time = data.get("close_time", "")
    if not open_time or not close_time:
        return jsonify(success=False, message="时间不能为空")
    try:
        open_dt = datetime.strptime(open_time, "%Y-%m-%dT%H:%M")
        close_dt = datetime.strptime(close_time, "%Y-%m-%dT%H:%M")
    except ValueError:
        return jsonify(success=False, message="时间格式错误")
    if close_dt <= open_dt:
        return jsonify(success=False, message="结束时间必须晚于开始时间")
    conn = get_db()
    conn.execute("UPDATE projects SET open_time=?, close_time=?, timer_enabled=1 WHERE id=?",
                 (open_dt.strftime("%Y-%m-%dT%H:%M"), close_dt.strftime("%Y-%m-%dT%H:%M"), pid))
    conn.commit()
    conn.close()
    return jsonify(success=True)

# ========== 班级管理 ==========
@app.route("/admin/project/<int:pid>/classes")
@admin_required
def admin_classes(pid):
    conn = get_db()
    classes = conn.execute("""
        SELECT c.*, (SELECT COUNT(*) FROM students s WHERE s.class_id=c.id AND s.project_id=?) as student_count
        FROM classes c WHERE c.project_id=? ORDER BY c.name
    """, (pid, pid)).fetchall()
    conn.close()
    return render_template("admin_classes.html", classes=classes, pid=pid)

@app.route("/admin/project/<int:pid>/classes/add", methods=["POST"])
@admin_required
def admin_classes_add(pid):
    name = request.form.get("name", "").strip()
    if name:
        conn = get_db()
        try:
            conn.execute("INSERT INTO classes (name, project_id) VALUES (?, ?)", (name, pid))
            conn.commit()
            flash(f"班级「{name}」添加成功", "success")
        except:
            flash("班级已存在", "danger")
        conn.close()
    return redirect(url_for("admin_classes", pid=pid))

@app.route("/admin/project/<int:pid>/classes/delete/<int:cid>")
@admin_required
def admin_classes_delete(pid, cid):
    conn = get_db()
    conn.execute("UPDATE students SET class_id=NULL WHERE class_id=? AND project_id=?", (cid, pid))
    conn.execute("DELETE FROM position_classes WHERE class_id=?", (cid,))
    conn.execute("DELETE FROM classes WHERE id=? AND project_id=?", (cid, pid))
    conn.commit()
    conn.close()
    flash("班级已删除", "warning")
    return redirect(url_for("admin_classes", pid=pid))

# ========== 学生管理 ==========
@app.route("/admin/project/<int:pid>/students")
@admin_required
def admin_students(pid):
    conn = get_db()
    classes = conn.execute("SELECT * FROM classes WHERE project_id=? ORDER BY name", (pid,)).fetchall()
    selected_class = request.args.get("class_id", "")
    if selected_class:
        students = conn.execute("""
            SELECT s.*, c.name as class_name FROM students s
            LEFT JOIN classes c ON s.class_id=c.id WHERE s.project_id=? AND s.class_id=?
            ORDER BY c.name, s.student_id
        """, (pid, selected_class)).fetchall()
    else:
        students = conn.execute("""
            SELECT s.*, c.name as class_name FROM students s
            LEFT JOIN classes c ON s.class_id=c.id WHERE s.project_id=?
            ORDER BY c.name, s.student_id
        """, (pid,)).fetchall()
    conn.close()
    return render_template("admin_students.html", students=students, classes=classes, 
                           selected_class=selected_class, pid=pid)

@app.route("/admin/project/<int:pid>/students/import", methods=["POST"])
@admin_required
def admin_students_import(pid):
    f = request.files.get("file")
    class_id = request.form.get("class_id")
    if not f or not class_id:
        flash("请选择文件和班级", "danger")
        return redirect(url_for("admin_students", pid=pid))
    try:
        wb = load_workbook(f)
        ws = wb.active
        conn = get_db()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                try:
                    conn.execute("INSERT OR IGNORE INTO students (name, student_id, class_id, project_id) VALUES (?, ?, ?, ?)",
                                (str(row[0]).strip(), str(row[1]).strip(), int(class_id), pid))
                    count += 1
                except:
                    pass
        conn.commit()
        conn.close()
        flash(f"成功导入 {count} 名学生", "success")
    except Exception as e:
        flash(f"导入失败: {e}", "danger")
    return redirect(url_for("admin_students", pid=pid))

@app.route("/admin/project/<int:pid>/students/template")
def student_template(pid):
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "学号"])
    ws.append(["张三", "2021001"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="student_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/project/<int:pid>/students/delete/<int:sid>")
@admin_required
def admin_students_delete(pid, sid):
    conn = get_db()
    s = conn.execute("SELECT * FROM students WHERE id=? AND project_id=?", (sid, pid)).fetchone()
    if s and s["has_submitted"]:
        conn.execute("UPDATE positions SET current_count=MAX(0,current_count-1) WHERE id IN (SELECT position_id FROM submissions WHERE student_id=? AND project_id=?)", (sid, pid))
    conn.execute("DELETE FROM submissions WHERE student_id=? AND project_id=?", (sid, pid))
    conn.execute("DELETE FROM student_position_assign WHERE student_id=? AND project_id=?", (sid, pid))
    conn.execute("DELETE FROM students WHERE id=? AND project_id=?", (sid, pid))
    conn.commit()
    conn.close()
    flash("学生已删除", "warning")
    return redirect(url_for("admin_students", pid=pid))

@app.route("/admin/project/<int:pid>/students/batch_delete", methods=["POST"])
@admin_required
def admin_students_batch_delete(pid):
    ids = request.form.get("ids", "")
    if not ids:
        flash("未选择学生", "warning")
        return redirect(url_for("admin_students", pid=pid))
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    conn = get_db()
    for sid in id_list:
        s = conn.execute("SELECT * FROM students WHERE id=? AND project_id=?", (sid, pid)).fetchone()
        if s and s["has_submitted"]:
            conn.execute("UPDATE positions SET current_count=MAX(0,current_count-1) WHERE id IN (SELECT position_id FROM submissions WHERE student_id=? AND project_id=?)", (sid, pid))
        conn.execute("DELETE FROM submissions WHERE student_id=? AND project_id=?", (sid, pid))
        conn.execute("DELETE FROM student_position_assign WHERE student_id=? AND project_id=?", (sid, pid))
        conn.execute("DELETE FROM students WHERE id=? AND project_id=?", (sid, pid))
    conn.commit()
    conn.close()
    flash(f"已删除 {len(id_list)} 名学生", "success")
    return redirect(url_for("admin_students", pid=pid))

# ========== 岗位管理 ==========
@app.route("/admin/project/<int:pid>/positions")
@admin_required
def admin_positions(pid):
    conn = get_db()
    positions = conn.execute("""
        SELECT p.*, 
               GROUP_CONCAT(c.name) as class_names,
               GROUP_CONCAT(pc.class_id) as visible_class_ids
        FROM positions p
        LEFT JOIN position_classes pc ON p.id = pc.position_id
        LEFT JOIN classes c ON pc.class_id = c.id
        WHERE p.project_id = ?
        GROUP BY p.id ORDER BY p.id
    """, (pid,)).fetchall()
    classes = conn.execute("SELECT * FROM classes WHERE project_id=? ORDER BY name", (pid,)).fetchall()
    conn.close()
    return render_template("admin_positions.html", positions=positions, classes=classes, pid=pid)

@app.route("/admin/project/<int:pid>/positions/add", methods=["POST"])
@admin_required
def admin_positions_add(pid):
    base = request.form.get("base_name", "").strip()
    quota = int(request.form.get("quota", 1))
    instructor = request.form.get("instructor", "").strip()
    requirements = request.form.get("requirements", "").strip()
    class_ids = request.form.getlist("class_ids")
    if not base:
        flash("请填写实习基地名称", "danger")
        return redirect(url_for("admin_positions", pid=pid))
    conn = get_db()
    cur = conn.execute("INSERT INTO positions (base_name, quota, instructor, requirements, project_id) VALUES (?, ?, ?, ?, ?)",
                       (base, quota, instructor, requirements, pid))
    pos_id = cur.lastrowid
    for cid in class_ids:
        conn.execute("INSERT OR IGNORE INTO position_classes (position_id, class_id) VALUES (?, ?)", (pos_id, int(cid)))
    conn.commit()
    conn.close()
    flash(f"岗位「{base}」添加成功", "success")
    return redirect(url_for("admin_positions", pid=pid))

@app.route("/admin/project/<int:pid>/positions/edit/<int:pos_id>", methods=["POST"])
@admin_required
def admin_positions_edit(pid, pos_id):
    base = request.form.get("base_name", "").strip()
    quota = int(request.form.get("quota", 1))
    instructor = request.form.get("instructor", "").strip()
    requirements = request.form.get("requirements", "").strip()
    class_ids = request.form.getlist("class_ids")
    conn = get_db()
    conn.execute("UPDATE positions SET base_name=?, quota=?, instructor=?, requirements=? WHERE id=? AND project_id=?",
                 (base, quota, instructor, requirements, pos_id, pid))
    conn.execute("DELETE FROM position_classes WHERE position_id=?", (pos_id,))
    for cid in class_ids:
        conn.execute("INSERT OR IGNORE INTO position_classes (position_id, class_id) VALUES (?, ?)", (pos_id, int(cid)))
    conn.commit()
    conn.close()
    flash("岗位已更新", "success")
    return redirect(url_for("admin_positions", pid=pid))

@app.route("/admin/project/<int:pid>/positions/delete/<int:pos_id>")
@admin_required
def admin_positions_delete(pid, pos_id):
    conn = get_db()
    conn.execute("UPDATE students SET has_submitted=0, submitted_at=NULL, assigned_position_id=NULL WHERE id IN (SELECT student_id FROM submissions WHERE position_id=? AND project_id=?)", (pos_id, pid))
    conn.execute("DELETE FROM submissions WHERE position_id=? AND project_id=?", (pos_id, pid))
    conn.execute("DELETE FROM student_position_assign WHERE position_id=? AND project_id=?", (pos_id, pid))
    conn.execute("DELETE FROM position_classes WHERE position_id=?", (pos_id,))
    conn.execute("DELETE FROM positions WHERE id=? AND project_id=?", (pos_id, pid))
    conn.commit()
    conn.close()
    flash("岗位已删除", "warning")
    return redirect(url_for("admin_positions", pid=pid))

@app.route("/admin/project/<int:pid>/positions/batch_delete", methods=["POST"])
@admin_required
def admin_positions_batch_delete(pid):
    ids = request.form.get("ids", "")
    if not ids:
        flash("未选择岗位", "warning")
        return redirect(url_for("admin_positions", pid=pid))
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    conn = get_db()
    for pos_id in id_list:
        conn.execute("UPDATE students SET has_submitted=0, submitted_at=NULL, assigned_position_id=NULL WHERE id IN (SELECT student_id FROM submissions WHERE position_id=? AND project_id=?)", (pos_id, pid))
        conn.execute("DELETE FROM submissions WHERE position_id=? AND project_id=?", (pos_id, pid))
        conn.execute("DELETE FROM student_position_assign WHERE position_id=? AND project_id=?", (pos_id, pid))
        conn.execute("DELETE FROM position_classes WHERE position_id=?", (pos_id,))
        conn.execute("DELETE FROM positions WHERE id=? AND project_id=?", (pos_id, pid))
    conn.commit()
    conn.close()
    flash(f"已删除 {len(id_list)} 个岗位", "success")
    return redirect(url_for("admin_positions", pid=pid))

@app.route("/admin/project/<int:pid>/positions/import", methods=["POST"])
@admin_required
def admin_positions_import(pid):
    f = request.files.get("file")
    if not f:
        flash("请选择文件", "danger")
        return redirect(url_for("admin_positions", pid=pid))
    try:
        wb = load_workbook(f)
        ws = wb.active
        conn = get_db()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            base = str(row[0]).strip() if row[0] else ""
            quota = int(row[1]) if row[1] else 1
            instructor = str(row[2]).strip() if row[2] else ""
            req = str(row[3]).strip() if row[3] else ""
            class_names = str(row[4]).strip() if row[4] else ""
            if base:
                cur = conn.execute("INSERT INTO positions (base_name, quota, instructor, requirements, project_id) VALUES (?,?,?,?,?)",
                                   (base, quota, instructor, req, pid))
                pos_id = cur.lastrowid
                if class_names:
                    for cn in class_names.split(","):
                        cn = cn.strip()
                        cr = conn.execute("SELECT id FROM classes WHERE name=? AND project_id=?", (cn, pid)).fetchone()
                        if cr:
                            conn.execute("INSERT OR IGNORE INTO position_classes (position_id, class_id) VALUES (?,?)", (pos_id, cr["id"]))
                count += 1
        conn.commit()
        conn.close()
        flash(f"成功导入 {count} 个岗位", "success")
    except Exception as e:
        flash(f"导入失败: {e}", "danger")
    return redirect(url_for("admin_positions", pid=pid))

@app.route("/admin/project/<int:pid>/positions/template")
def position_template(pid):
    wb = Workbook()
    ws = wb.active
    ws.append(["实习基地", "配额人数", "指导教师", "备注要求", "可见班级(逗号分隔)"])
    ws.append(["XX医院", 5, "张老师", "需持护士证", "护理1班,护理2班"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="position_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========== 指定关系 ==========
@app.route("/admin/project/<int:pid>/assign")
@admin_required
def admin_assign(pid):
    conn = get_db()
    students = conn.execute("""
        SELECT s.*, c.name as class_name FROM students s
        LEFT JOIN classes c ON s.class_id=c.id WHERE s.project_id=? ORDER BY c.name, s.student_id
    """, (pid,)).fetchall()
    positions = conn.execute("SELECT * FROM positions WHERE project_id=? ORDER BY base_name", (pid,)).fetchall()
    assigns = conn.execute("""
        SELECT spa.student_id, spa.position_id, s.name as sname, s.student_id as sid, p.base_name
        FROM student_position_assign spa
        JOIN students s ON spa.student_id=s.id
        JOIN positions p ON spa.position_id=p.id
        WHERE spa.project_id=?
        ORDER BY s.student_id
    """, (pid,)).fetchall()
    conn.close()
    return render_template("admin_assign.html", students=students, positions=positions, assigns=assigns, pid=pid)

@app.route("/admin/project/<int:pid>/assign/add", methods=["POST"])
@admin_required
def admin_assign_add(pid):
    sid = request.form.get("student_id")
    pos_id = request.form.get("position_id")
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO student_position_assign (student_id, position_id, project_id) VALUES (?, ?, ?)", 
                 (int(sid), int(pos_id), pid))
    conn.commit()
    conn.close()
    flash("指定成功", "success")
    return redirect(url_for("admin_assign", pid=pid))

@app.route("/admin/project/<int:pid>/assign/delete/<int:sid>")
@admin_required
def admin_assign_delete(pid, sid):
    conn = get_db()
    conn.execute("DELETE FROM student_position_assign WHERE student_id=? AND project_id=?", (sid, pid))
    conn.commit()
    conn.close()
    flash("已取消指定", "warning")
    return redirect(url_for("admin_assign", pid=pid))

@app.route("/admin/project/<int:pid>/assign/import", methods=["POST"])
@admin_required
def admin_assign_import(pid):
    f = request.files.get("file")
    if not f:
        flash("请选择文件", "danger")
        return redirect(url_for("admin_assign", pid=pid))
    try:
        wb = load_workbook(f)
        ws = wb.active
        conn = get_db()
        count = 0
        errors = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            sid = str(row[0]).strip() if row[0] else ""
            pname = str(row[1]).strip() if row[1] else ""
            if not sid or not pname:
                continue
            student = conn.execute("SELECT id FROM students WHERE student_id=? AND project_id=?", (sid, pid)).fetchone()
            position = conn.execute("SELECT id FROM positions WHERE base_name=? AND project_id=?", (pname, pid)).fetchone()
            if not student:
                errors.append(f"第{i}行：学号「{sid}」不存在")
                continue
            if not position:
                errors.append(f"第{i}行：岗位「{pname}」不存在")
                continue
            conn.execute("INSERT OR REPLACE INTO student_position_assign (student_id, position_id, project_id) VALUES (?, ?, ?)",
                        (student["id"], position["id"], pid))
            count += 1
        conn.commit()
        conn.close()
        if count:
            flash(f"成功指定 {count} 条记录", "success")
        if errors:
            flash("；".join(errors[:10]), "warning")
    except Exception as e:
        flash(f"导入失败: {e}", "danger")
    return redirect(url_for("admin_assign", pid=pid))

@app.route("/admin/project/<int:pid>/assign/template")
def assign_template(pid):
    wb = Workbook()
    ws = wb.active
    ws.append(["学号", "岗位名称"])
    ws.append(["2021001", "XX医院"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="assign_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========== 提交记录 ==========
@app.route("/admin/project/<int:pid>/submissions")
@admin_required
def admin_submissions(pid):
    conn = get_db()
    rows = conn.execute("""
        SELECT s.name, s.student_id, c.name as class_name, p.base_name, sub.created_at
        FROM submissions sub
        JOIN students s ON sub.student_id=s.id
        LEFT JOIN classes c ON s.class_id=c.id
        JOIN positions p ON sub.position_id=p.id
        WHERE sub.project_id=?
        ORDER BY sub.created_at DESC
    """, (pid,)).fetchall()
    conn.close()
    return render_template("admin_submissions.html", rows=rows, pid=pid)

# ========== 导出 ==========
@app.route("/admin/project/<int:pid>/export")
@admin_required
def admin_export(pid):
    conn = get_db()
    rows = conn.execute("""
        SELECT s.student_id, s.name, c.name as class_name,
               p.base_name, p.instructor, p.requirements, sub.created_at
        FROM submissions sub
        JOIN students s ON sub.student_id=s.id
        LEFT JOIN classes c ON s.class_id=c.id
        JOIN positions p ON sub.position_id=p.id
        WHERE sub.project_id=?
        ORDER BY c.name, s.student_id
    """, (pid,)).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "选择结果"
    ws.append(["学号", "姓名", "班级", "实习基地", "指导教师", "备注要求", "提交时间"])
    for r in rows:
        ws.append([r["student_id"], r["name"], r["class_name"], r["base_name"], r["instructor"], r["requirements"], r["created_at"]])
    ws2 = wb.create_sheet("岗位统计")
    ws2.append(["实习基地", "配额", "已选", "剩余"])
    stats = conn.execute("SELECT base_name, quota, current_count FROM positions WHERE project_id=? ORDER BY base_name", (pid,)).fetchall()
    for s in stats:
        ws2.append([s["base_name"], s["quota"], s["current_count"], s["quota"] - s["current_count"]])
    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    project = conn.execute("SELECT name FROM projects WHERE id=?", (pid,)).fetchone() if not conn.isolation_level else None
    fname = f"选岗结果_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========== Student API ==========
@app.route("/api/system_status/<int:pid>")
def api_system_status(pid):
    conn = get_db()
    row = conn.execute("SELECT status FROM projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    mode = row["status"] if row else "close"
    return jsonify({
        "mode": mode,
        "open": mode == "open",
        "preview": mode == "preview",
        "can_view": mode in ("preview", "open")
    })

@app.route("/api/positions/<int:pid>/<int:student_id>")
def api_positions(pid, student_id):
    conn = get_db()
    row = conn.execute("SELECT status FROM projects WHERE id=?", (pid,)).fetchone()
    if not row or row["status"] == "close":
        conn.close()
        return jsonify([])
    
    student = conn.execute("SELECT id, class_id FROM students WHERE id=? AND project_id=?", (student_id, pid)).fetchone()
    if not student:
        conn.close()
        return jsonify([])
    
    student_db_id = student["id"]
    class_id = student["class_id"]
    random_order = conn.execute("SELECT value FROM settings WHERE key='random_order' AND project_id=?", (pid,)).fetchone()
    is_random = random_order and random_order["value"] == "1"
    
    existing = conn.execute("SELECT position_id FROM submissions WHERE student_id=? AND project_id=?", (student_db_id, pid)).fetchone()
    submitted_position_id = existing["position_id"] if existing else None
    
    if class_id:
        positions = conn.execute("""
            SELECT p.id, p.base_name, p.quota, p.current_count, p.instructor, p.requirements
            FROM positions p
            JOIN position_classes pc ON p.id = pc.position_id
            WHERE pc.class_id = ? AND p.project_id = ?
        """, (class_id, pid)).fetchall()
    else:
        positions = conn.execute("SELECT id, base_name, quota, current_count, instructor, requirements FROM positions WHERE project_id=?", (pid,)).fetchall()
    
    shared = conn.execute("""
        SELECT p.id, p.base_name, p.quota, p.current_count, p.instructor, p.requirements
        FROM positions p
        WHERE p.project_id=? AND p.id NOT IN (SELECT DISTINCT position_id FROM position_classes)
    """, (pid,)).fetchall()
    
    my_assigned = set()
    rows = conn.execute("SELECT position_id FROM student_position_assign WHERE student_id=? AND project_id=?", (student_db_id, pid)).fetchall()
    my_assigned = set(r["position_id"] for r in rows)
    
    all_positions_raw = {}
    for p in list(positions) + list(shared):
        if p["id"] not in all_positions_raw:
            all_positions_raw[p["id"]] = dict(p)
    
    assigned_counts = {}
    for pos_id in all_positions_raw.keys():
        row = conn.execute("SELECT COUNT(*) as cnt FROM student_position_assign WHERE position_id=? AND project_id=?", (pos_id, pid)).fetchone()
        assigned_counts[pos_id] = row["cnt"] if row else 0
    
    # 获取项目开放时间
    proj = conn.execute("SELECT open_time FROM projects WHERE id=?", (pid,)).fetchone()
    open_minutes = -1
    if proj and proj["open_time"]:
        try:
            open_dt = datetime.strptime(proj["open_time"], "%Y-%m-%dT%H:%M")
            open_minutes = (datetime.now() - open_dt).total_seconds() / 60
        except:
            pass
    
    assigned_list = []
    normal_list = []
    
    for pos_id, p in all_positions_raw.items():
        assigned_cnt = assigned_counts.get(pos_id, 0)
        is_my = pos_id in my_assigned
        
        if is_my:
            remaining = max(0, p["quota"] - p["current_count"])
        else:
            occupied_designated = 0
            if assigned_cnt > 0:
                designated_students = conn.execute(
                    "SELECT student_id FROM student_position_assign WHERE position_id=? AND project_id=?", (pos_id, pid)
                ).fetchall()
                for ds in designated_students:
                    sub = conn.execute(
                        "SELECT id FROM submissions WHERE student_id=? AND position_id=? AND project_id=?",
                        (ds["student_id"], pos_id, pid)
                    ).fetchone()
                    if sub:
                        occupied_designated += 1
                    elif open_minutes >= 1:
                        occupied_designated += 1
            remaining = max(0, p["quota"] - occupied_designated - p["current_count"])
        
        is_selected = submitted_position_id == pos_id
        
        item = {
            "id": pos_id,
            "base_name": p["base_name"],
            "quota": p["quota"],
            "remaining": remaining,
            "instructor": p["instructor"],
            "requirements": p["requirements"],
            "assigned": is_my,
            "selected": is_selected
        }
        
        if is_my:
            assigned_list.append(item)
        else:
            normal_list.append(item)
    
    if is_random and student_db_id:
        random.Random(student_db_id).shuffle(normal_list)
    
    conn.close()
    return jsonify(assigned_list + normal_list)

# ========== Student Pages ==========
@app.route("/")
def index():
    return redirect(url_for("student_login"))

@app.route("/student/login", methods=["GET", "POST"])
def student_login():
    conn = get_db()
    projects = conn.execute("SELECT id, name, status FROM projects WHERE status != 'draft' ORDER BY id").fetchall()
    conn.close()
    
    if request.method == "POST":
        pid = request.form.get("project_id")
        name = request.form.get("name", "").strip()
        sid = request.form.get("student_id", "").strip()
        if not pid:
            flash("请选择项目", "danger")
            return render_template("student_login.html", projects=projects)
        
        pid = int(pid)
        check_time_based_open(pid)
        
        conn = get_db()
        student = conn.execute("""
            SELECT s.*, c.name as class_name FROM students s 
            LEFT JOIN classes c ON s.class_id=c.id 
            WHERE s.name=? AND s.student_id=? AND s.project_id=?
        """, (name, sid, pid)).fetchone()
        
        if student:
            session["student_id"] = student["id"]
            session["student_name"] = student["name"]
            session["student_sid"] = student["student_id"]
            session["class_id"] = student["class_id"]
            session["class_name"] = student["class_name"]
            session["project_id"] = pid
            
            proj = conn.execute("SELECT status, name FROM projects WHERE id=?", (pid,)).fetchone()
            conn.close()
            
            if student["has_submitted"]:
                session.clear()
                flash("你已完成岗位选择", "info")
                return redirect(url_for("student_login"))
            if proj["status"] == "close":
                session.pop("student_id", None)
                flash("系统暂未开放", "warning")
                return redirect(url_for("student_login"))
            return redirect(url_for("student_select"))
        conn.close()
        flash("姓名或学号不正确", "danger")
    return render_template("student_login.html", projects=projects)

@app.route("/student/select", methods=["GET", "POST"])
def student_select():
    if "student_id" not in session or "project_id" not in session:
        return redirect(url_for("student_login"))
    
    pid = session["project_id"]
    check_time_based_open(pid)
    
    conn = get_db()
    proj = conn.execute("SELECT status, name FROM projects WHERE id=?", (pid,)).fetchone()
    mode = proj["status"] if proj else "close"
    
    if mode == "close":
        conn.close()
        flash("系统暂未开放", "warning")
        return redirect(url_for("student_login"))
    
    student = conn.execute("SELECT * FROM students WHERE id=? AND project_id=?", (session["student_id"], pid)).fetchone()
    if student["has_submitted"]:
        conn.close()
        return redirect(url_for("student_result"))
    
    assigned = conn.execute("""
        SELECT p.* FROM student_position_assign spa
        JOIN positions p ON spa.position_id=p.id
        WHERE spa.student_id=? AND spa.project_id=?
    """, (session["student_id"], pid)).fetchone()
    conn.close()
    
    if request.method == "POST":
        if mode != "open":
            flash("当前为预览模式，暂不能选择", "warning")
            return redirect(url_for("student_select"))
        pos_id = request.form.get("position_id")
        if not pos_id:
            flash("请选择一个岗位", "danger")
            return redirect(url_for("student_select"))
        pos_id = int(pos_id)
        conn = get_db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            pos = conn.execute("SELECT * FROM positions WHERE id=? AND project_id=?", (pos_id, pid)).fetchone()
            if not pos:
                conn.rollback()
                flash("岗位不存在", "danger")
                conn.close()
                return redirect(url_for("student_select"))
            is_my_assigned = conn.execute(
                "SELECT 1 FROM student_position_assign WHERE position_id=? AND student_id=? AND project_id=?",
                (pos_id, session["student_id"], pid)
            ).fetchone()
            assigned_count = conn.execute(
                "SELECT COUNT(*) as cnt FROM student_position_assign WHERE position_id=? AND project_id=?",
                (pos_id, pid)
            ).fetchone()["cnt"]
            if is_my_assigned:
                if pos["current_count"] >= pos["quota"]:
                    conn.rollback()
                    flash("该岗位名额已满", "danger")
                    conn.close()
                    return redirect(url_for("student_select"))
            else:
                effective_quota = max(0, pos["quota"] - assigned_count)
                if pos["current_count"] >= effective_quota:
                    conn.rollback()
                    flash("该岗位名额已满，请选择其他岗位", "danger")
                    conn.close()
                    return redirect(url_for("student_select"))
            conn.execute("UPDATE positions SET current_count=current_count+1 WHERE id=? AND current_count<quota", (pos_id,))
            conn.execute("INSERT INTO submissions (student_id, position_id, project_id) VALUES (?, ?, ?)",
                        (session["student_id"], pos_id, pid))
            conn.execute("UPDATE students SET has_submitted=1, submitted_at=datetime('now','localtime'), assigned_position_id=? WHERE id=? AND project_id=?",
                        (pos_id, session["student_id"], pid))
            conn.commit()
            conn.close()
            return redirect(url_for("student_result"))
        except Exception as e:
            conn.rollback()
            conn.close()
            flash(f"选择失败: {e}", "danger")
            return redirect(url_for("student_select"))
    
    return render_template("student_select.html",
                           student_id=session.get("student_id"),
                           class_id=session.get("class_id"),
                           class_name=session.get("class_name"),
                           student_name=session.get("student_name"),
                           project_id=pid,
                           project_name=proj["name"] if proj else "",
                           system_mode=mode)

@app.route("/student/result")
def student_result():
    if "student_id" not in session or "project_id" not in session:
        return redirect(url_for("student_login"))
    pid = session["project_id"]
    conn = get_db()
    sub = conn.execute("""
        SELECT p.base_name, p.instructor, p.requirements, sub.created_at
        FROM submissions sub JOIN positions p ON sub.position_id=p.id
        WHERE sub.student_id=? AND sub.project_id=?
    """, (session["student_id"], pid)).fetchone()
    proj = conn.execute("SELECT name FROM projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    if not sub:
        return redirect(url_for("student_login"))
    return render_template("student_result.html", sub=sub,
                           student_name=session.get("student_name"),
                           class_name=session.get("class_name"),
                           project_name=proj["name"] if proj else "")

@app.route("/student/logout")
def student_logout():
    session.clear()
    return redirect(url_for("student_login"))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)